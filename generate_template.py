import openpyxl
import os

# Paths
accurate_path = '/home/hasanarofid/Downloads/daftar-barang(1).xlsx'
price_list_path = '/home/hasanarofid/Downloads/Price List Mr Lux.xlsx'
output_path = '/home/hasanarofid/Documents/hasanarofid/project-agung/luxindonesia/@luxindonesialaravel/public/dokumen/template-import-produk.xlsx'

# 1. Read Price List and create a mapping of Name -> Price to try and populate missing prices
price_mapping = {}
if os.path.exists(price_list_path):
    wb_price = openpyxl.load_workbook(price_list_path, data_only=True)
    for sheet_name in wb_price.sheetnames:
        ws = wb_price[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 4: # Assuming Name is col 0 or 4, Price is col 3 roughly
                # Price list structure varies, let's try to extract best we can
                # Often Name is in col 0 (if "EPOXY..."), Price is in col 3
                # Or for Lem Putih, Name is in col 0, Price is in col 2
                name = str(row[0]).strip() if row[0] else ''
                
                # Check column 2, 3, 4 for a number that looks like a price
                price = 0
                for v in row[2:5]:
                    if isinstance(v, (int, float)) and v > 1000:
                        price = v
                        break
                if name and price:
                    price_mapping[name.lower()] = price

# 2. Read Accurate Export and build the template
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Daftar Produk"

headers = [
    'Nama', 'Kategori', 'SKU', 'Satuan', 
    'Isi per Dus', 'Isi per Set', 'Harga per Unit', 
    'Harga per Dus', 'Harga per Set', 'Total Stok', 'Deskripsi'
]
ws_out.append(headers)

if os.path.exists(accurate_path):
    wb_acc = openpyxl.load_workbook(accurate_path, data_only=True)
    # the exact data we saw previously
    for sheet_name in wb_acc.sheetnames:
        ws_acc = wb_acc[sheet_name]
        # Skip header row (assuming row 1 is header)
        for idx, row in enumerate(ws_acc.iter_rows(min_row=2, values_only=True)):
            # Based on previous analysis of daftar-barang(1).xlsx:
            # col 1 = 'Kategori Barang' (0-indexed: row[1])
            # col 2 = 'No. Barang' (SKU) (row[2])
            # col 3 = 'Keterangan' (Name) (row[3])
            # col 5 = 'Satuan 1' (row[5])
            # col 7 = 'Kuantitas 1' (Isi per Dus often) (row[7])
            # col 23 = 'Harga Jual 1' (row[23])
            
            if not row or len(row) < 24:
                continue
                
            kategori = row[1] if row[1] else ''
            sku = row[2] if row[2] else ''
            nama = row[3] if row[3] else ''
            
            if not sku and not nama:
                continue
                
            satuan = row[5] if row[5] else 'PCS'
            
            # Try to get isi/dus
            isi_dus = 1
            if row[7] and isinstance(row[7], (int, float, str)):
                try:
                    isi_dus = float(row[7])
                except:
                    pass
            
            # Try to get price from Accurate first
            harga = 0
            if row[23] and isinstance(row[23], (int, float, str)):
                try:
                    harga = float(row[23])
                except:
                    pass
            
            # If Accurate price is 0, try to find in Price List
            if harga == 0 and nama:
                n_lower = str(nama).lower()
                # Try exact match
                if n_lower in price_mapping:
                    harga = price_mapping[n_lower]
                else:
                    # Try partial match
                    for pn, pprice in price_mapping.items():
                        if pn in n_lower or n_lower in pn:
                            if pprice > harga:
                                harga = pprice
            
            out_row = [
                nama,        # Nama
                kategori,    # Kategori
                sku,         # SKU/Kode
                satuan,      # Satuan
                isi_dus,     # Isi per Dus
                1,           # Isi per Set
                harga,       # Harga per Unit
                0,           # Harga per Dus
                0,           # Harga per Set
                0,           # Total Stok
                ''           # Deskripsi
            ]
            ws_out.append(out_row)

wb_out.save(output_path)
print(f"Successfully generated template at: {output_path}")
